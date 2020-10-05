import sys

from pandas import DataFrame, Index, concat, to_datetime, to_timedelta, DatetimeIndex, ExcelWriter, Series
from pandas.io.formats.excel import ExcelFormatter
from process import resample_noise
from openpyxl import load_workbook
from datetime import datetime
from itertools import repeat


def daily_table(data, f_weight, max_remove, lmax_override=None):

    """

    Returns list of six tables summarising user-input time periods on a daily basis:

        1. main                 All metrics non-spectral summary    Aggregated by day
        2. main_summary         All metrics non-spectral summary    Aggregated over entire duration of data
        3. spec_leq_main        Leq spectral summary                Aggregated by day
        4. spec_leq_summary     Leq spectral summary                Aggregated over entire duration of data
        5. spec_lmax_main       Lmax spectral summary               Aggregated by day
        6. spec_lmax_summary    Lmax spectral summary               Aggregated over entire duration of data

    data = input DataFrame
    periods = dictionary where keys are names of time periods, values are START times
    leq, lmax, l90 = relevant column names in data

    Linear average used for Leq in main_summary based on feedback:

        [31/03/20 11:45 AM] Stephen Scott
        ...the average day/night should be a linear average of the whole-day/night periods. I think you've done the log.
        So a log average of, say the day's LAeq,5minutes will give a LAeq,16hr metric.
        For the whole survey over many days, one still wants the LAeq,16hr metric, so a linear average preserves that.

    """

    # Get list of time periods created with process.flag_periods()
    flags = data.filter(regex='^Flag_').columns.to_list()

    # Extract relevant columns from data
    leq = 'L' + f_weight + 'eq_Main'
    lmax = 'L' + f_weight + 'max_Main'
    l10 = 'L' + f_weight + '10_Main'
    l90 = 'L' + f_weight + '90_Main'
    leq_spectra = data.filter(regex='L' + f_weight + 'eq_.*Hz').columns.to_list()

    main_metrics = []
    for m in [leq, lmax, l10, l90]:
        if m in data.columns:
            main_metrics.append(m)

    df = data[main_metrics + leq_spectra + flags].copy()

    df_out = DataFrame()

    if not flags:
        df['Flag_24hr_Day'] = datetime.strptime(
            df.index.min().floor('1D').strftime('%d/%m/%y %H:%M'), '%d/%m/%y %H:%M'
        )
        flags = ['Flag_24hr_Day']

    flags_copy = flags.copy()
    for f in flags:

        # Filter data to period
        df_tmp = df.dropna(subset=[f]).copy()

        if df_tmp.empty:
            flags_copy.remove(f)

        else:

            df_tmp['Period'] = f.replace('Flag_', '')

            df_tmp['Start_Time'] = df_tmp.index.copy()
            df_tmp['End_Time'] = df_tmp.index.copy()

            # Shift data by period's start time to account for overnight cases
            t_start = to_timedelta(df_tmp[f][0].time().strftime('%H:%M:%S'))
            df_tmp.index = df_tmp.index - t_start

            # Use process.resample_noise() function to get daily summary
            df_mean = resample_noise(df_tmp, ['1D', max_remove, 'mean', f_weight, [10, 90], 'log'])[0]

            if (l10 in df_tmp.columns) and (l90 in df_tmp.columns):

                # Repeat to obtain mode and lower quartile for L10/L90 only
                df_mode = resample_noise(df_tmp[[l10, l90]], ['1D', max_remove, 'mode', f_weight, [10, 90], 'log'])[0]
                df_lq = resample_noise(df_tmp[[l10, l90]], ['1D', max_remove, 'lq', f_weight, [10, 90], 'log'])[0]
    
                # Merge mode and LQ with general (mean) summary
                df_tmp = df_mean.merge(
                    df_mode,
                    left_index=True,
                    right_index=True,
                    how='outer',
                    suffixes=['_mean', '']
                )
    
                df_tmp = df_tmp.merge(
                    df_lq,
                    left_index=True,
                    right_index=True,
                    how='outer',
                    suffixes=['_mode', '_lq']
                )

            else:
                df_tmp = df_mean

            # Invert time-shift
            df_tmp.index = df_tmp.index + t_start

            # Combine with previous time periods
            df_out = concat([df_out, df_tmp])

    flags = flags_copy

    # Set up aggregation method by column
    cols_out = {
        'Day': 'count',
        'Period': 'count',
        'Start_Time': 'mode',
        'End_Time': 'mode',
        leq: 'mean',
    }

    if lmax in df.columns:
        cols_out[lmax] = 'mean'
    if l10 in df.columns:
        cols_out[l10 + '_mean'] = 'mean'
        cols_out[l10 + '_mode'] = 'mean'
        cols_out[l10 + '_lq'] = 'mean'
    if l90 in df.columns:
        cols_out[l90 + '_mean'] = 'mean'
        cols_out[l90 + '_mode'] = 'mean'
        cols_out[l90 + '_lq'] = 'mean'

    for c in leq_spectra:
        cols_out[c] = 'mean'

    # Sort table
    df_out = df_out.sort_values('Start_Time')

    # Re-format days/times
    df_out['Day'] = df_out.index.weekday_name
    df_out['Start_Time'] = df_out['Start_Time'].dt.strftime('%H:%M')
    df_out['End_Time'] = df_out['End_Time'].dt.strftime('%H:%M')

    df_out = df_out[cols_out.keys()]
    df_out.index = DatetimeIndex(df_out.index.date, name='Date')
    df_out.dropna(inplace=True)

    # Derive all-time summary from daily table
    df_tmp = df_out.copy()

    cols_out_max = {}
    cols_out_mode = {}
    for c in cols_out:
        if cols_out[c] == 'mean':
            cols_out_max[c] = 'max'
            cols_out_mode[c] = 'mode'
        else:
            cols_out_max[c] = cols_out[c]
            cols_out_mode[c] = cols_out[c]

    df_mean = DataFrame(columns=cols_out.keys())
    df_max = DataFrame(columns=cols_out.keys())
    df_mode = DataFrame(columns=cols_out.keys())

    for i, f in enumerate(flags):

        f = f.replace('Flag_', '')

        df_mean.loc[i] = df_tmp[df_tmp['Period'] == f].agg(cols_out).loc[0]
        df_max.loc[i] = df_tmp[df_tmp['Period'] == f].agg(cols_out_max).loc[0]
        df_mode.loc[i] = df_tmp[df_tmp['Period'] == f].round().agg(cols_out_mode).min()

        df_mean.loc[i, 'Period'] = f
        df_max.loc[i, 'Period'] = f
        df_mode.loc[i, 'Period'] = f

        df_mean.loc[i, 'Day'] = 'Mean (full survey)'
        df_max.loc[i, 'Day'] = 'Max (full survey)'
        df_mode.loc[i, 'Day'] = 'Lowest Mode (full survey)'

        df_mode.loc[i, 'Start_Time'] = df_mean.loc[i, 'Start_Time']
        df_mode.loc[i, 'End_Time'] = df_mean.loc[i, 'End_Time']

    # Collate tables
    df_main_cols = df_out.drop(columns=leq_spectra).columns
    df_spec_cols = ['Day', 'Period', 'Start_Time', 'End_Time'] + leq_spectra + ['L' + f_weight + 'eq_Main']

    df_out.index = df_out.index.date

    main = df_out[df_main_cols]
    main_mean = df_mean[df_main_cols]
    main_max = df_max[df_main_cols]
    main_mode = df_mode[df_main_cols]

    # Catch non-spectral data
    spectral = False
    for c in data.columns:
        if 'Hz' in c:
            spectral = True
            break

    if spectral:

        spec_leq_main = df_out[df_spec_cols]
        spec_leq_mean = df_mean[df_spec_cols]
        spec_leq_max = df_max[df_spec_cols]
        spec_leq_mode = df_mode[df_spec_cols]

        if lmax in data.columns:
            spec_lmax_main = lmax_spectra(data, main, f_weight, flags, summary=False)
            spec_lmax_mean = lmax_spectra(data, main_mean, f_weight, flags, summary=True)
            spec_lmax_max = lmax_spectra(data, main_max, f_weight, flags, summary=True)
            spec_lmax_mode = lmax_spectra(data, main_mode, f_weight, flags, summary=True)

            if lmax_override:
                spec_lmax_user = lmax_spectra(data, main_mean, f_weight, flags, summary=True, lmax_override=lmax_override)
                spec_lmax_user['Day'] = 'User-defined Lmax (full survey)'
            else:
                spec_lmax_user = DataFrame(index=Index([], name="No user-defined Lmax"))

        else:
            spec_lmax_main = DataFrame(index=Index([], name="No Lmax data"))
            spec_lmax_mean = DataFrame(index=Index([], name="No Lmax data"))
            spec_lmax_max = DataFrame(index=Index([], name="No Lmax data"))
            spec_lmax_mode = DataFrame(index=Index([], name="No Lmax data"))
            spec_lmax_user = DataFrame(index=Index([], name="No Lmax data"))

    else:

        spec_leq_main = DataFrame(index=Index([], name="No spectral data available"))
        spec_lmax_main = DataFrame(index=Index([], name="No spectral data available"))
        spec_leq_mean, spec_leq_max, spec_leq_mode, spec_lmax_mean, spec_lmax_max, spec_lmax_mode, spec_lmax_user = \
            repeat(DataFrame(), 7)

    tables = [
        main, main_mean, main_max, main_mode,
        spec_leq_main, spec_leq_mean, spec_leq_max, spec_leq_mode,
        spec_lmax_main, spec_lmax_mean, spec_lmax_max, spec_lmax_mode, spec_lmax_user
    ]

    # Add one sample to End Time column (assumes fixed sample rate)
    for i, t in enumerate(tables):
        if 'End_Time' in t.columns:

            end_time = (to_datetime(t.loc[:, 'End_Time']) + data.index.freq).apply(lambda x: x.strftime('%H:%M'))
            t = t.drop(columns='End_Time')
            t.insert(3, 'End_Time', end_time)
            tables[i] = t

    return tables


def lmax_spectra(data, table, f_weight, flags, summary=False, lmax_override={}):

    """

    Generates representative Lmax summary table from daily_table() outputs.
    Uses least squares minimisation on narrow band of 125-4000 Hz to find the sample with spectrum closest to the mean.
    Called from daily_table() function.

    :param data         : Full dataset, used to look up all instances of selected Lmax
    :param table        : 'Main' output table from outputs.daily_table(), either first or second element
    :param f_weight     : Frequency weighting (string), e.g. 'A'
    :param flags        : List of time periods created with process.flag_periods(), e.g.
                                ['Flag_Daytime', 'Flag_Night-time']
    :param summary      : True for overall summary (tables[0]); False for daily summary (tables[1])
    :param lmax_override: Dict of integer values with which to override automatic Lmax (11th value), e.g.
                                {'Daytime': 67, 'Night-time': 54}

    :return             : Table in the same format as daily_tables(), with representative Lmax spectra

    TODO: verify approach leading to discrepancy between main here and Lmax summary from daily_table()
            (rounding Lmax_Main to nearest dB prior to lookup)

    """

    lmax = 'L' + f_weight + 'max_Main'
    lmax_cols = data.filter(regex='max').columns.to_list()

    # Extract columns relating to 125-4000 Hz
    narrow_cols = []
    for c in lmax_cols:
        if ('Hz' in c.split('_')) and (125 <= float(c.split('_')[c.split('_').index('Hz') - 1]) <= 4000):
            narrow_cols.append(c)

    df_out = DataFrame()

    for f in flags:

        # Filter data and summary table to period
        df_tmp = data.dropna(subset=[f])[flags + lmax_cols].copy()
        tab_tmp = table[table['Period'] == f.replace('Flag_', '')][[
            'Day',
            'Period',
            'Start_Time',
            'End_Time',
            'L' + f_weight + 'max_Main'
        ]].copy()

        # Manual override
        f_replace = f.replace('Flag_', '')
        if summary and (f_replace in lmax_override.keys()):
            tab_tmp['L' + f_weight + 'max_Main'] = lmax_override[f_replace]

        df_tmp[lmax + '_rnd'] = df_tmp[lmax].round(0)
        tab_tmp[lmax + '_rnd'] = tab_tmp[lmax].round(0)

        merge_left = [lmax + '_rnd']
        merge_right = [lmax + '_rnd']

        if summary:

            tab_tmp['day'] = 0

        else:

            tab_tmp.index = tab_tmp.index.astype('<M8[ns]')

            # Shift data by period's start time to account for overnight cases
            t_start = to_timedelta(df_tmp[f][0].time().strftime('%H:%M:%S'))
            df_tmp.index = df_tmp.index - t_start

            # Prepare columns for matching
            tab_tmp['day'] = tab_tmp.index.date
            df_tmp['day'] = df_tmp.index.date
            merge_left.append('day')
            merge_right.append('day')

        # Look up instances of each day's Lmax in filtered data
        tab_tmp = tab_tmp.merge(
            df_tmp,
            left_on=merge_left,
            right_on=merge_right,
            how='left',
            suffixes=['_table', '']
        ).fillna(value={'day': 0}).drop(columns=[lmax + '_rnd'])

        agg = tab_tmp[['day'] + narrow_cols].groupby('day').mean()
        tab_tmp = tab_tmp.merge(
            agg,
            left_on='day',
            right_index=True,
            suffixes=['', '_tmp'],
            how='left'
        )

        for c in narrow_cols:
            tab_tmp[c + '_tmp'] = (tab_tmp[c] - tab_tmp[c + '_tmp'])**2

        tab_tmp['square_sum'] = tab_tmp.filter(regex='_tmp').sum(axis=1)
        agg = tab_tmp[['day', 'square_sum']].groupby('day').min().reset_index()
        tab_tmp = tab_tmp.merge(
            agg,
            on=['day', 'square_sum'],
            how='inner'
        ).drop(columns=[s + '_tmp' for s in narrow_cols] + ['square_sum'])

        # Combine with previous time periods
        df_out = concat([df_out, tab_tmp])

    if not summary:
        # Sort table
        df_out.index = to_datetime(df_out['day'].astype(str) + ' ' + df_out['Start_Time'])
        df_out.sort_index(inplace=True)
        df_out.index = df_out.index.date

    lmax_cols.sort(key=('L' + f_weight + 'max_Main').__eq__)
    df_spec_cols = ['Day', 'Period', 'Start_Time', 'End_Time'] + lmax_cols

    return df_out[df_spec_cols]     # .round(0)


def export_excel(data, metadata, tables, config):

    """

    Export full data and results summary tables to Excel template

    Note on template file:
        Ideally, pivot tables would read full data (i.e. un-rounded) and band with a grouping of 0.5-99.5 at 1dB inc.
        Upon testing, it was found that openpyxl.load_workbook is unable to read grouped fields in pivot tables
        Solution will be to apply banding within Python function if possible
        If not possible, solution will be to include extra columns, rounded to 0 d.p.

    """

    file_out = config["output"][0] + ".xlsm"
    config_out = file_out.replace(".xlsm", "_config.txt")

    # Create summary
    flags = data.filter(regex='^Flag_').columns.to_list()
    s1 = Series(
        index=[
            'Number of Time Samples',
            'Number of Recorded Metrics',
            'Number of User-defined Time Periods',
        ],
        data=[
            len(data),
            len(data.columns) - len(flags),
            len(flags),
        ]
    )

    s2 = Series(
        index=['Time Period ' + str(f + 1) for f in range(len(flags))],
        data=flags
    ).str.replace('Flag_', '')

    s3 = Series(
        index=[
            '',
            'Monitor Metadata:'
        ],
        data=[
            None,
            None
        ]
    )

    summary = concat([s1, s2, s3, metadata])

    # Re-order columns
    cols = ['Address', 'Duration']
    cols += data.filter(regex='_Main$').columns.to_list()
    cols += data.filter(regex='Hz$').columns.to_list()
    cols += flags
    for c in data.columns:
        if c not in cols:
            cols += [c]
    data_out = data[cols].copy()

    # Replace flags' start times with booleans
    for f in flags:
        data_out_tmp = data_out[f].dropna()
        if data_out_tmp.empty:
            data_out[f] = False
        else:
            t_start = data_out_tmp[0]
            data_out[f] = data_out[f].replace({t_start: True, None: False})

    # Collate tables

    for i, t in enumerate(tables):
        if i not in list(range(12))[0::4]:
            t.index = len(t) * ['']

    table_main = concat([
        tables[0],
        DataFrame(columns=tables[0].columns, index=['']),
        tables[1],
        DataFrame(columns=tables[0].columns, index=['']),
        tables[2],
        DataFrame(columns=tables[0].columns, index=['']),
        tables[3]
    ])

    table_leq_spec = concat([
        tables[4],
        DataFrame(columns=tables[4].columns, index=['']),
        tables[5],
        DataFrame(columns=tables[4].columns, index=['']),
        tables[6],
        DataFrame(columns=tables[4].columns, index=['']),
        tables[7]
    ])

    table_lmax_spec = concat([
        tables[8],
        DataFrame(columns=tables[8].columns, index=['']),
        tables[9],
        DataFrame(columns=tables[8].columns, index=['']),
        tables[10],
        DataFrame(columns=tables[8].columns, index=['']),
        tables[11],
        DataFrame(columns=tables[8].columns, index=['']),
        tables[12]
    ])

    # Read template
    file_template = "OutputExcelTemplate_v09_20200922.xlsm"
    wb = load_workbook(file_template, keep_vba=True)

    # Write full data
    ExcelFormatter.header_style = None
    writer = ExcelWriter(
        file_out,
        engine='openpyxl',
        datetime_format='dd/mm/yyyy hh:mm:ss',
        date_format='dd/mm/yyyy'
    )
    writer.book = wb
    writer.sheets = dict((ws.title, ws) for ws in wb.worksheets)

    print("Exporting to Excel...")

    # Write data
    summary.to_excel(writer, "Summary", header=False)
    data_out.to_excel(writer, "Full_Data")
    table_main.to_excel(writer, "Summary_Tables")
    table_leq_spec.to_excel(writer, "Leq_Spectral_Tables")
    table_lmax_spec.to_excel(writer, "Lmax_Spectral_Tables")

    return writer, config_out
